# views.py
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.db.models import Avg, Sum, F, Q, DecimalField
from django.db.models.functions import Coalesce
from django.db import transaction
from .models import *
from .forms import *
from openpyxl import load_workbook, Workbook
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from io import BytesIO
from datetime import datetime, timedelta
from datetime import date
import calendar
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.db.models import Sum, Avg, F, DecimalField
from django.db.models.functions import Coalesce
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from datetime import datetime, date
import calendar
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Sum, Avg, F, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import datetime
import calendar
from django.db.models import Count

from .models import ProductType, Product, CartProducts, ProductPrice

from django.contrib.auth.decorators import login_required
import io
import base64
import matplotlib
matplotlib.use('Agg') 
from django import forms
from django.shortcuts import render
# myapp/views.py
from django.shortcuts import render
from rest_framework import generics
from .models import Commune, Wilaya, PointOfSale, Product, ProductPrice
from .serializers import CommuneGeoJSONSerializer,MoughataaGeoJSONSerializer, WilayaGeoJSONSerializer, PointOfSaleSerializer, ProductSerializer, ProductPriceSerializer
from django.db.models import Avg, Sum, F, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import datetime
from django.http import JsonResponse
import json

# myapp/views.py
from rest_framework.views import APIView
from rest_framework.response import Response

class MoughataaGeoList(APIView):
    def get(self, request):
        moughataas = Moughataa.objects.all()
        serializer = MoughataaGeoJSONSerializer(moughataas, many=True)
        geojson = {
            "type": "FeatureCollection",
            "features": serializer.data,
        }
        return Response(geojson)
    
class WilayaGeoList(APIView):
    def get(self, request):
        wilayas = Wilaya.objects.all()
        features = WilayaGeoJSONSerializer(wilayas, many=True).data
        geojson = {
            "type": "FeatureCollection",
            "features": features
        }
        return Response(geojson)

class CommuneGeoList(APIView):
    def get(self, request):
        communes = Commune.objects.all()
        features = CommuneGeoJSONSerializer(communes, many=True).data
        geojson = {
            "type": "FeatureCollection",
            "features": features
        }
        return Response(geojson)
    
class PointOfSaleList(generics.ListAPIView):
    queryset = PointOfSale.objects.all()
    serializer_class = PointOfSaleSerializer

class ProductList(generics.ListAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

class ProductPriceList(generics.ListAPIView):
    queryset = ProductPrice.objects.all()
    serializer_class = ProductPriceSerializer


from django.http import JsonResponse

from decimal import Decimal
from django.db.models import Sum, F, DecimalField
from django.db.models.functions import Coalesce
from datetime import date

# Global calculation over all CartProducts (for the given month)
def calculer_inpc_global_mensuel(annee, mois):
    # Filter CartProducts where the linked price's date is in the specified month/year
    cart_products_qs = CartProducts.objects.filter(
        price__date__year=annee,
        price__date__month=mois
    )
    
    # Sum of the weights
    somme_ponderations = cart_products_qs.aggregate(
        total_pond=Coalesce(Sum('weight'), Decimal('0.00'))
    )['total_pond']
    
    # Sum of (price * weight)
    somme_produits_ponderees = cart_products_qs.aggregate(
        total_pondere=Coalesce(
            Sum(F('price__value') * F('weight'), output_field=DecimalField()),
            Decimal('0.00')
        )
    )['total_pondere']
    
    inpc = (somme_produits_ponderees / somme_ponderations * Decimal('100')) if somme_ponderations > Decimal('0.00') else Decimal('0.00')
    return inpc

# Calculation per wilaya using the modified relationship
def calculer_inpc_wilaya(annee, mois, wilaya_code):
    # Filter CartProducts where:
    #  - the linked price's date is in the given month/year, and
    #  - the price's point of sale is in a commune that belongs to a moughataa whose wilaya has the given code.
    cart_products_qs = CartProducts.objects.filter(
        price__date__year=annee,
        price__date__month=mois,
        price__point_of_sale__commune__moughataa__wilaya__code=wilaya_code
    )
    
    somme_ponderations = cart_products_qs.aggregate(
        total_pond=Coalesce(Sum('weight'), Decimal('0.00'))
    )['total_pond']
    
    somme_produits_ponderees = cart_products_qs.aggregate(
        total_pondere=Coalesce(
            Sum(F('price__value') * F('weight'), output_field=DecimalField()),
            Decimal('0.00')
        )
    )['total_pondere']
    
    inpc = (somme_produits_ponderees / somme_ponderations * Decimal('100')) if somme_ponderations > Decimal('0.00') else Decimal('0.00')
    return inpc

# (Optional) Calculation per commune – similar idea:
def calculer_inpc_commune(annee, mois, commune_code):
    cart_products_qs = CartProducts.objects.filter(
        price__date__year=annee,
        price__date__month=mois,
        price__point_of_sale__commune__code=commune_code
    )
    
    somme_ponderations = cart_products_qs.aggregate(
        total_pond=Coalesce(Sum('weight'), Decimal('0.00'))
    )['total_pond']
    
    somme_produits_ponderees = cart_products_qs.aggregate(
        total_pondere=Coalesce(
            Sum(F('price__value') * F('weight'), output_field=DecimalField()),
            Decimal('0.00')
        )
    )['total_pondere']
    
    inpc = (somme_produits_ponderees / somme_ponderations * Decimal('100')) if somme_ponderations > Decimal('0.00') else Decimal('0.00')
    return inpc

from django.http import JsonResponse
from datetime import date

def calculate_inpc(request):
    # For example, use the last 6 months period.
    # Here, we'll compute the INPC over the last 6 months and then distribute it by wilaya.
    # For simplicity, let’s use a fixed period for testing (say January 2025).
    annee = 2025
    mois = 1

    # First, compute the global INPC over this month (or 6-month period if you adapt the function)
    global_inpc = calculer_inpc_global_mensuel(annee, mois)
    print(f"Global INPC for {annee}-{mois}: {global_inpc}")

    # Now update each Wilaya with its INPC value (for the given period)
    wilayas = Wilaya.objects.all()
    updates = []
    for wilaya in wilayas:
        wilaya_inpc = calculer_inpc_wilaya(annee, mois, wilaya.code)
        wilaya.inpc = wilaya_inpc
        wilaya.save()
        updates.append({'wilaya': wilaya.name, 'inpc': str(wilaya_inpc)})
        print(f"Updated Wilaya {wilaya.name} (code: {wilaya.code}) to INPC: {wilaya_inpc}")
    
    return JsonResponse({
        'message': 'INPC for last 6 months calculated and distributed to Wilayas',
        'global_inpc': str(global_inpc),
        'updates': updates
    })


def mrmap_view(request):
    return render(request, 'mrmap.html')

def debug_geometries(request):
    wilayas_debug = Wilaya.objects.all().values('name', 'code', 'polygon')[:5]
    communes_debug = Commune.objects.all().values('name', 'code', 'polygon')[:5]
    
    return JsonResponse({
        'wilayas': list(wilayas_debug),
        'communes': list(communes_debug)
    })
# ====================
# Chart Generation
# ====================
def get_inpc_chart_data():
    today = datetime.now()
    last_6_months = []
    current_year = today.year
    current_month = today.month

    for i in range(6):
        mo = current_month - i
        yr = current_year
        if mo <= 0:
            mo += 12
            yr -= 1
        last_6_months.append((yr, mo))

    inpc_data = []
    for (yr, mo) in reversed(last_6_months):
        inpc_value = calculer_inpc_global_mensuel(annee=yr, mois=mo)
        inpc_data.append({
            'year': yr,
            'month': mo,
            'month_name': calendar.month_name[mo],
            'inpc': inpc_value
        })

    return inpc_data
def calculer_inpc_global_mensuel(annee, mois):
    """
    Calcule l'Indice National des Prix à la Consommation (INPC) pour un mois et une année donnés.

    Args:
        annee (int): L'année pour laquelle calculer l'INPC.
        mois (int): Le mois pour lequel calculer l'INPC.

    Returns:
        Decimal: La valeur de l'INPC calculée.
    """
    prix_entries = ProductPrice.objects.filter(
        date_from__year=annee,
        date_from__month=mois
    )

    # Calcul de la somme des pondérations via la relation produit -> CartProducts
    somme_ponderations = prix_entries.aggregate(
        total_pond=Coalesce(Sum('product__cartproducts__weight'), Decimal('0.00'))
    )['total_pond']

    # Calcul de la somme des produits pondérés (prix * pondération)
    somme_produits_ponderees = prix_entries.aggregate(
        total_pondere=Coalesce(
            Sum(F('value') * F('product__cartproducts__weight'), output_field=DecimalField()),
            Decimal('0.00')
        )
    )['total_pondere']

    # Calcul de l'INPC en pourcentage
    inpc = (somme_produits_ponderees / somme_ponderations * Decimal('100')) if somme_ponderations > 0 else Decimal('0.00')
    return inpc
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Count, Sum, Avg
from datetime import datetime
import calendar
from decimal import Decimal
from django.db.models.functions import Coalesce
from django.db.models import DecimalField

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Avg
from datetime import datetime
import calendar
from decimal import Decimal
from django.core.serializers.json import DjangoJSONEncoder
import json

@login_required
def home_view(request):
    """
    Génère le tableau de bord interactif avec les statistiques et graphiques Django-ChartJS.
    """
    today = datetime.now()

    # 🔹 Générer la liste des 6 derniers mois
    last_6_months = [(today.year, today.month - i) for i in range(6)]
    last_6_months = [(yr - 1, mo + 12) if mo <= 0 else (yr, mo) for yr, mo in last_6_months]
    
    inpc_data = []
    for (yr, mo) in reversed(last_6_months):
        inpc_value = calculer_inpc_global_mensuel(annee=yr, mois=mo)
        inpc_data.append({
            'year': yr,
            'month': mo,
            'month_name': calendar.month_name[mo],
            'inpc': round(inpc_value, 2)
        })

    # 🔹 Calcul des statistiques globales
    total_products_calculated = Product.objects.count()
    total_weights = CartProducts.objects.aggregate(total_weight=Sum('weight'))['total_weight'] or Decimal('0.00')
    inpc_global = calculer_inpc_global_mensuel(today.year, today.month) or Decimal('0.00')
    average_inpc_6_months = sum(item['inpc'] for item in inpc_data) / len(inpc_data) if inpc_data else Decimal('0.00')

    # 📊 Préparer les données pour les graphiques
    products_by_type = Product.objects.values('product_type__label').annotate(count=Count('id'))
    product_type_distribution = ProductType.objects.annotate(num_products=Count('product'))
    avg_price_by_product_type = ProductPrice.objects.values('product__product_type__label').annotate(avg_price=Avg('value'))
    pos_by_commune = PointOfSale.objects.values('commune__name').annotate(count=Count('id'))
    total_weight_by_cart = CartProducts.objects.values('cart__name').annotate(total_weight=Sum('weight'))

    chart_data = {
        'inpc_trends': {
            'labels': [item['month_name'] for item in inpc_data],
            'data': [item['inpc'] for item in inpc_data],
        },
        'products_by_type': {
            'labels': [item['product_type__label'] for item in products_by_type],
            'data': [item['count'] for item in products_by_type],
        },
        'product_type_distribution': {
            'labels': [pt.label for pt in product_type_distribution],
            'data': [pt.num_products for pt in product_type_distribution],
        },
        'avg_price_by_product_type': {
            'labels': [item['product__product_type__label'] for item in avg_price_by_product_type],
            'data': [round(item['avg_price'], 2) if item['avg_price'] else 0 for item in avg_price_by_product_type],
        },
        'pos_by_commune': {
            'labels': [item['commune__name'] for item in pos_by_commune],
            'data': [item['count'] for item in pos_by_commune],
        },
        'total_weight_by_cart': {
            'labels': [item['cart__name'] for item in total_weight_by_cart],
            'data': [round(item['total_weight'], 2) if item['total_weight'] else 0 for item in total_weight_by_cart],
        },
    }

    # 🔹 Sérialiser les données des graphiques en JSON
    chart_data_json = json.dumps(chart_data, cls=DjangoJSONEncoder)

    # 🔹 Rendre le template avec les données statistiques et graphiques
    context = {
        'inpc_data': inpc_data,
        'inpc_global': round(inpc_global, 2),
        'total_products_calculated': total_products_calculated,
        'total_weights': round(total_weights, 2),
        'average_inpc': round(average_inpc_6_months, 2),
        'chart_data': chart_data_json,  # Ajouter les données des graphiques
    }
    # Add empty forms for modals
    context['producttype_form'] = ProductTypeForm()
    context['product_form'] = ProductForm()
    context['productprice_form'] = ProductPriceForm()
    context['cart_form'] = CartForm()
    context['cartproducts_form'] = CartProductsForm()
    context['pointofsale_form'] = PointOfSaleForm()
    return render(request, 'home.html', context)


@login_required
def commune_list_view(request):
    """
    Vue unique gérant à la fois :
      - L'affichage HTML (filtrage, recherche, pagination)
      - Les appels AJAX (renvoie un JSON filtré/paginé)
      - Filtrage par wilaya / moughataa
      - Recherche par nom ou code
      - Export en Excel (format .xlsx)
    """

    # --- 1) Lecture des paramètres GET ---
    selected_wilaya = request.GET.get('wilaya', '').strip()
    selected_moughataa = request.GET.get('moughataa', '').strip()
    search = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', 1)

    # Paramètre pour export Excel, ex.: ?export=excel
    export_param = request.GET.get('export', '').strip()

    # --- 2) Construction du queryset (filtrage) ---
    communes_qs = Commune.objects.all()

    if selected_wilaya:
        communes_qs = communes_qs.filter(moughataa__wilaya__code=selected_wilaya)

    if selected_moughataa:
        communes_qs = communes_qs.filter(moughataa__code=selected_moughataa)

    if search:
        communes_qs = communes_qs.filter(
            Q(name__icontains=search) | Q(code__icontains=search)
        )

    # --- 3) Export en Excel (si ?export=excel) ---
    if export_param == 'excel':
        # Crée un classeur Excel en mémoire
        wb = Workbook()
        ws = wb.active
        ws.title = "Communes"

        # En-tête
        ws.append(["ID", "Code", "Nom", "Moughataa", "Wilaya"])

        # Ajout des lignes : pas de pagination pour l'export
        for commune in communes_qs:
            ws.append([
                commune.id,
                commune.code,
                commune.name,
                commune.moughataa.name,
                commune.moughataa.wilaya.name
            ])

        # Conversion en bytes pour la réponse HTTP
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=communes.xlsx"
        response.write(output.getvalue())
        return response

    # --- 4) Sinon, on poursuit la logique standard (HTML, AJAX) ---

    # Pagination
    paginator = Paginator(communes_qs, 10)  # 10 communes par page
    page_obj = paginator.get_page(page_number)

    # Moughataas dynamiques (pour le <select>)
    if selected_wilaya:
        moughataas_qs = Moughataa.objects.filter(wilaya__code=selected_wilaya)
    else:
        moughataas_qs = Moughataa.objects.all()

    # Mode AJAX ?
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            "communes": [],
            "pagination": {
                "current_page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        }
        for commune in page_obj:
            data["communes"].append({
                "id": commune.id,
                "code": commune.code,
                "name": commune.name,
                "moughataa": commune.moughataa.name,
                "wilaya": commune.moughataa.wilaya.name
            })
        return JsonResponse(data, safe=False)

    # Mode HTML classique
    context = {
        'page_obj': page_obj,
        'wilayas': Wilaya.objects.all(),
        'moughataas': moughataas_qs,
        'selected_wilaya': selected_wilaya,
        'selected_moughataa': selected_moughataa,
        'search': search
    }
    return render(request, 'commune.html', context)

@login_required
def import_data_view(request):
    """
    Gère l'import des données Excel pour
    Wilaya, Moughataa, et Commune.
    """
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if excel_file:
            wb = load_workbook(excel_file)
            with transaction.atomic():
                # Feuille 1 : Wilayas
                sheet_wilayas = wb["mrt_adm_adm1_2024.shp"]
                for row in sheet_wilayas.iter_rows(min_row=2, values_only=True):
                    wilaya_name = row[2]  # ADM1_EN
                    wilaya_code = row[3]  # ADM1_PCODE
                    Wilaya.objects.update_or_create(
                        code=wilaya_code,
                        defaults={"name": wilaya_name}
                    )

                # Feuille 2 : Moughataas
                sheet_moughataas = wb["mrt_adm_adm2_2024.shp"]
                for row in sheet_moughataas.iter_rows(min_row=2, values_only=True):
                    wilaya_code = row[3]       # ADM1_PCODE
                    moughataa_name = row[4]    # ADM2_EN
                    moughataa_code = row[5]    # ADM2_PCODE
                    linked_wilaya = Wilaya.objects.get(code=wilaya_code)
                    Moughataa.objects.update_or_create(
                        code=moughataa_code,
                        defaults={
                            "name": moughataa_name,
                            "wilaya": linked_wilaya
                        }
                    )

                # Feuille 3 : Communes
                sheet_communes = wb["mrt_adm_adm3_2024.shp"]
                for row in sheet_communes.iter_rows(min_row=2, values_only=True):
                    moughataa_code = row[5]  # ADM2_PCODE
                    commune_name = row[6]    # ADM3_EN
                    commune_code = row[7]    # ADM3_PCODE
                    linked_moughataa = Moughataa.objects.get(code=moughataa_code)
                    Commune.objects.update_or_create(
                        code=commune_code,
                        defaults={
                            "name": commune_name,
                            "moughataa": linked_moughataa
                        }
                    )
            # Redirection vers la liste générale des communes après import
            return redirect("commune-list")

    # Méthode GET ou si pas de fichier, on peut juste rediriger vers la liste
    return redirect("commune-list")

@login_required
def point_of_sale_list_view(request):
    """
    Affiche la liste des points de vente (PointOfSale) avec :
      - Filtrage par commune, type
      - Recherche (code ou type)
      - Pagination
      - Réponse JSON (si requête AJAX)
      - Export Excel (via ?export=excel)
    """

    # 1) Paramètres GET
    selected_commune = request.GET.get('commune', '').strip()
    selected_type = request.GET.get('type', '').strip()
    search = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', 1)
    export_param = request.GET.get('export', '').strip()  # ?export=excel

    # 2) Construction du queryset (filtrage)
    pos_qs = PointOfSale.objects.all()

    if selected_commune:
        pos_qs = pos_qs.filter(commune__code=selected_commune)

    if selected_type:
        pos_qs = pos_qs.filter(type__iexact=selected_type)

    if search:
        pos_qs = pos_qs.filter(
            Q(code__icontains=search) | Q(type__icontains=search)
        )

    # 3) Export en Excel (si ?export=excel)
    if export_param == 'excel':
        # Créer le fichier Excel en mémoire
        wb = Workbook()
        ws = wb.active
        ws.title = "PointsOfSale"

        # En-têtes de colonnes
        ws.append(["ID", "Code", "Type", "GPS Lat", "GPS Lon", "Commune"])

        # On exporte l'ensemble du queryset (pas de pagination)
        for pos in pos_qs:
            commune_name = pos.commune.name if pos.commune else ""
            ws.append([
                pos.id,
                pos.code,
                pos.type,
                pos.gps_lat,
                pos.gps_lon,
                commune_name,
            ])

        # Conversion en bytes pour la réponse HTTP
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="points_of_sale.xlsx"'
        response.write(output.getvalue())
        return response

    # 4) Sinon, logique standard (HTML ou AJAX)
    paginator = Paginator(pos_qs, 10)  # 10 items par page
    page_obj = paginator.get_page(page_number)

    # Réponse AJAX (JSON)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            "points_of_sale": [],
            "pagination": {
                "current_page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        }
        for pos in page_obj.object_list:
            data["points_of_sale"].append({
                "id": pos.id,
                "code": pos.code,
                "type": pos.type,
                "gps_lat": pos.gps_lat,
                "gps_lon": pos.gps_lon,
                "commune": pos.commune.name if pos.commune else "",
            })
        return JsonResponse(data, safe=False)

    # Réponse HTML classique
    communes_list = Commune.objects.all().order_by("name")
    types_list = (PointOfSale.objects
                  .values_list("type", flat=True)
                  .distinct()
                  .order_by("type"))

    context = {
        "page_obj": page_obj,
        "communes": communes_list,
        "available_types": types_list,
        "selected_commune": selected_commune,
        "selected_type": selected_type,
        "search": search,
    }
    
    # Add empty forms for modals
    context['pointofsale_form'] = PointOfSaleForm()
    return render(request, "pointofsale_list.html", context)

@login_required
def import_point_of_sale_view(request):
    """
    Gère l'import Excel pour PointOfSale.
    Suppose qu'on lit un fichier avec colonnes :
      Code | Type | GPS_Lat | GPS_Lon | Commune_Code
    """
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect("pointofsale-list")

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active  # ou wb["SheetName"] si nécessaire

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue  # ligne vide

                    code = row[0]
                    pos_type = row[1]
                    gps_lat = row[2]
                    gps_lon = row[3]
                    commune_code = row[4]

                    # Vérification de la commune
                    commune_obj = None
                    if commune_code:
                        commune_obj = Commune.objects.get(code=commune_code)

                    # Mise à jour ou création
                    PointOfSale.objects.update_or_create(
                        code=code,
                        defaults={
                            "type": pos_type,
                            "gps_lat": gps_lat,
                            "gps_lon": gps_lon,
                            "commune": commune_obj
                        }
                    )
            messages.success(request, "Importation réussie.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")

    # Redirection vers la liste
    return redirect("pointofsale-list")

def point_of_sale_detail(request, pk):
    point_of_sale = PointOfSale.objects.get(pk=pk)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
       return render(request, 'pointofsale_detail_modal.html', {'point_of_sale': point_of_sale})
    
    return redirect('pointofsale-list')

def point_of_sale_create(request):
    if request.method == 'POST':
        form = PointOfSaleForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Point de vente créé avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = PointOfSaleForm()
        return render(request, 'pointofsale_form_modal.html', {'form': form, 'type': 'create'})

    return redirect('pointofsale-list')


def point_of_sale_update(request, pk):
    point_of_sale = PointOfSale.objects.get(pk=pk)
    if request.method == 'POST':
        form = PointOfSaleForm(request.POST, instance=point_of_sale)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Point de vente mis à jour avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = PointOfSaleForm(instance=point_of_sale)
        return render(request, 'pointofsale_form_modal.html', {'form': form, 'type': 'update', 'point_of_sale': point_of_sale})

    return redirect('pointofsale-list')


def point_of_sale_delete(request, pk):
    point_of_sale = PointOfSale.objects.get(pk=pk)
    if request.method == 'POST':
        point_of_sale.delete()
        return JsonResponse({'success': True, 'message': 'Point de vente supprimé avec succès'})
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'pointofsale_confirm_delete_modal.html', {'point_of_sale': point_of_sale})
    return redirect('pointofsale-list')


# ====================
# Product Views
# ====================
@login_required
def product_list_view(request):
    """
    Affiche la liste des produits (Product) avec :
      - Filtrage par product_type
      - Recherche (code, name, product_type)
      - Pagination
      - Export Excel (via ?export=excel)
      - Réponse JSON (si requête AJAX)
    """

    # 1) Paramètres GET
    selected_type = request.GET.get('product_type', '').strip()
    search = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', 1)
    export_param = request.GET.get('export', '').strip()  # ?export=excel

    # 2) Construction du queryset (filtrage)
    products_qs = Product.objects.all()

    if selected_type:
        products_qs = products_qs.filter(product_type__iexact=selected_type)

    if search:
        # Recherche sur code, name, product_type
        products_qs = products_qs.filter(
            Q(code__icontains=search) |
            Q(name__icontains=search) |
            Q(product_type__icontains=search)
        )

    # 3) Export en Excel (si ?export=excel)
    if export_param == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Entêtes de colonnes
        ws.append(["ID", "Code", "Name", "Description", "Unit Measure", "Product Type"])

        # Pas de pagination pour l'export : on exporte TOUT le queryset filtré
        for prod in products_qs:
            ws.append([
                prod.id,
                prod.code,
                prod.name,
                prod.description,
                prod.unit_measure,
                prod.product_type
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
        response.write(output.getvalue())
        return response

    # 4) Mode normal (HTML ou AJAX)
    paginator = Paginator(products_qs, 10)  # 10 produits par page
    page_obj = paginator.get_page(page_number)

    # Mode AJAX => JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            "products": [],
            "pagination": {
                "current_page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        }
        for product in page_obj.object_list:
            data["products"].append({
                "id": product.id,
                "code": product.code,
                "name": product.name,
                "description": product.description,
                "unit_measure": product.unit_measure,
                "product_type": product.product_type.label,
            })
        return JsonResponse(data, safe=False)

    # Liste des types disponibles (distinct)
    product_types_list = (Product.objects
                          .values_list("product_type__label", flat=True)
                          .distinct()
                          .order_by("product_type__label"))

    context = {
        "page_obj": page_obj,
        "product_types": product_types_list,
        "selected_type": selected_type,
        "search": search
    }
    # Add empty forms for modals
    context['product_form'] = ProductForm()
    return render(request, "product_list.html", context)
@login_required
def import_product_view(request):
    """
    Gère l'import Excel pour Product.
    Suppose qu'on lit un fichier avec colonnes :
      code | name | description | unit_measure | product_type_code
      (ligne d'en-tête ignorée, min_row=2)
    """
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect("product-list")

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active  # ou wb["NomDeFeuille"] si besoin

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue  # ligne vide
                    code = row[0]
                    name = row[1]
                    description = row[2]
                    unit_measure = row[3]
                    product_type_code = row[4]  # Assurez-vous que c'est le code, pas le label

                    # Récupérer le ProductType correspondant
                    try:
                        product_type = ProductType.objects.get(code=product_type_code)
                    except ProductType.DoesNotExist:
                        messages.error(request, f"Type de produit '{product_type_code}' introuvable pour le produit '{code}'.")
                        continue  # Aller à la ligne suivante

                    # Mise à jour ou création
                    Product.objects.update_or_create(
                        code=code,
                        defaults={
                            "name": name,
                            "description": description,
                            "unit_measure": unit_measure,
                            "product_type": product_type  # Assigner l'objet ProductType
                        }
                    )
            messages.success(request, "Importation réussie.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")

    return redirect("product-list")

def product_detail(request, pk):
    product = Product.objects.get(pk=pk)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'product_detail_modal.html', {'product': product})
    return redirect('product-list')

def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Produit créé avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = ProductForm()
        return render(request, 'product_form_modal.html', {'form': form, 'type': 'create'})

    return redirect('product-list')


def product_update(request, pk):
    product = Product.objects.get(pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Produit mis à jour avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = ProductForm(instance=product)
        return render(request, 'product_form_modal.html', {'form': form, 'type': 'update', 'product': product})

    return redirect('product-list')

def product_delete(request, pk):
    product = Product.objects.get(pk=pk)
    if request.method == 'POST':
        product.delete()
        return JsonResponse({'success': True, 'message': 'Produit supprimé avec succès'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'product_confirm_delete_modal.html', {'product': product})

    return redirect('product-list')

# ====================
# Product Price Views
# ====================
# views.py (continued)

# ... (previous code)

# ====================
# Product Price Views
# ====================
@login_required
def productprice_list_view(request):
    """
    Affiche la liste des prix des produits (ProductPrice) avec :
      - Filtrage (ex. par product, par point_of_sale)
      - Recherche
      - Export Excel (si ?export=excel)
      - Pagination
      - Réponse JSON (si AJAX)
    """

    # 1) Lecture des paramètres GET
    selected_product = request.GET.get("product", "").strip()
    selected_pos = request.GET.get("point_of_sale", "").strip()
    search = request.GET.get("search", "").strip()
    page_number = request.GET.get("page", 1)
    export_param = request.GET.get("export", "").strip()  # ?export=excel

    # 2) Construction du queryset (filtrage)
    pp_qs = ProductPrice.objects.select_related("product", "point_of_sale").all()

    # Filtrage par produit (selon ID ou code)
    if selected_product:
        # si vous avez un champ "code" ou "id" => ajustez
        pp_qs = pp_qs.filter(product__id=selected_product)

    # Filtrage par point de vente
    if selected_pos:
        pp_qs = pp_qs.filter(point_of_sale__id=selected_pos)

    # Recherche globale
    if search:
        # Filtre sur product.name, point_of_sale.type ou code, etc.
        pp_qs = pp_qs.filter(
            Q(product__name__icontains=search) |
            Q(point_of_sale__type__icontains=search)
        )

    # 3) Export Excel si ?export=excel
    if export_param == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "ProductPrices"

        # En-têtes de colonnes
        ws.append(["ID", "Produit", "Point de Vente", "Prix", "Date Début", "Date Fin"])

        # Pas de pagination pour l'export
        for pp in pp_qs:
            product_name = pp.product.name if pp.product else ""
            pos_name = pp.point_of_sale.type if pp.point_of_sale else ""
            date_from = pp.date_from.strftime("%Y-%m-%d") if pp.date_from else ""
            date_to = pp.date_to.strftime("%Y-%m-%d") if pp.date_to else ""

            ws.append([
                pp.id,
                product_name,
                pos_name,
                pp.value,
                date_from,
                date_to,
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="product_prices.xlsx"'
        response.write(output.getvalue())
        return response

    # 4) Sinon, logique standard (HTML / AJAX)
    paginator = Paginator(pp_qs, 10)
    page_obj = paginator.get_page(page_number)

    # Mode AJAX => JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            "product_prices": [],
            "pagination": {
                "current_page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        }
        for pp in page_obj:
            data["product_prices"].append({
                "id": pp.id,
                "product_name": pp.product.name if pp.product else "",
                "point_of_sale": pp.point_of_sale.type if pp.point_of_sale else "",
                "value": pp.value,
                "date_from": pp.date_from.isoformat() if pp.date_from else None,
                "date_to": pp.date_to.isoformat() if pp.date_to else None
            })
        return JsonResponse(data, safe=False)

    # Mode HTML
    context = {
        "page_obj": page_obj,
        "search": search,
        "selected_product": selected_product,
        "selected_pos": selected_pos
    }
    # Add empty forms for modals
    context['productprice_form'] = ProductPriceForm()
    return render(request, "productprice_list.html", context)
@login_required
def import_productprice_view(request):
    """
    Gère l'import Excel pour ProductPrice.
    Suppose qu'on lit un fichier avec colonnes :
      product_code | pointofsale_code | value | date_from | date_to
    (adapter selon votre vrai fichier)
    """

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect("productprice-list")

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active  # ou wb["SheetName"] si besoin

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue

                    product_code = row[0]
                    pos_code = row[1]
                    value_price = row[2]
                    date_from_val = row[3]
                    date_to_val = row[4]

                    # Récupérer le Product
                    product_obj = None
                    if product_code:
                        product_obj = Product.objects.get(code=product_code)

                    # Récupérer le PointOfSale
                    pos_obj = None
                    if pos_code:
                        pos_obj = PointOfSale.objects.get(code=pos_code)

                    # Mise à jour / création
                    # Hypothèse : un ProductPrice est identifié par (product, point_of_sale, date_from ?)
                    ProductPrice.objects.update_or_create(
                        product=product_obj,
                        point_of_sale=pos_obj,
                        date_from=date_from_val,
                        defaults={
                            "value": value_price,
                            "date_to": date_to_val
                        }
                    )
            messages.success(request, "Importation réussie.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")

    return redirect("productprice-list")

def productprice_detail(request, pk):
    productprice = ProductPrice.objects.get(pk=pk)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'productprice_detail_modal.html', {'productprice': productprice})

    return redirect('productprice-list')

def productprice_create(request):
    if request.method == 'POST':
        form = ProductPriceForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Prix de produit créé avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = ProductPriceForm()
        return render(request, 'productprice_form_modal.html', {'form': form, 'type': 'create'})

    return redirect('productprice-list')

def productprice_update(request, pk):
    productprice = ProductPrice.objects.get(pk=pk)
    if request.method == 'POST':
        form = ProductPriceForm(request.POST, instance=productprice)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Prix de produit mis à jour avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = ProductPriceForm(instance=productprice)
        return render(request, 'productprice_form_modal.html', {'form': form, 'type': 'update', 'productprice': productprice})

    return redirect('productprice-list')

def productprice_delete(request, pk):
    productprice = ProductPrice.objects.get(pk=pk)
    if request.method == 'POST':
        productprice.delete()
        return JsonResponse({'success': True, 'message': 'Prix de produit supprimé avec succès'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'productprice_confirm_delete_modal.html', {'productprice': productprice})

    return redirect('productprice-list')

# ====================
# Cart Views
# ====================
@login_required
def cart_list_view(request):
    """
    Affiche la liste des paniers (Cart) avec :
      - Filtrage (exemple : recherche par code ou nom)
      - Export Excel (si ?export=excel)
      - Pagination
      - Réponse JSON (si AJAX)
    """

    # 1) Lecture des paramètres GET
    search = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', 1)
    export_param = request.GET.get('export', '').strip()  # ex: ?export=excel

    # 2) Construction du queryset
    carts_qs = Cart.objects.all()

    # Filtrage (recherche simple sur code ou nom)
    if search:
        carts_qs = carts_qs.filter(
            Q(code__icontains=search) |
            Q(name__icontains=search)
        )

    # 3) Export Excel
    if export_param == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Carts"

        # En-têtes
        ws.append(["ID", "Code", "Name", "Description"])

        # Export complet, sans pagination
        for cart in carts_qs:
            ws.append([
                cart.id,
                cart.code,
                cart.name,
                cart.description
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="carts.xlsx"'
        response.write(output.getvalue())
        return response

    # 4) Logique standard (HTML ou AJAX)
    paginator = Paginator(carts_qs, 10)  # 10 paniers par page
    page_obj = paginator.get_page(page_number)

    # Mode AJAX => renvoyer JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            "carts": [],
            "pagination": {
                "current_page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        }
        for cart in page_obj:
            data["carts"].append({
                "id": cart.id,
                "code": cart.code,
                "name": cart.name,
                "description": cart.description
            })
        return JsonResponse(data, safe=False)

    # Mode HTML classique
    context = {
        'carts': page_obj,     # vous pouvez nommer 'carts' ou 'page_obj' selon votre template
        'search': search
    }
     # Add empty forms for modals
    context['cart_form'] = CartForm()
    return render(request, 'cart_list.html', context)
@login_required
def import_cart_view(request):
    """
    Gère l'import Excel pour Cart.
    Suppose que le fichier Excel a des colonnes :
      code | name | description
      (Ligne d'en-tête ignorée, on commence à la ligne 2)
    """
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect("cart-list")

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active  # ou wb["SheetName"] si vous avez une feuille nommée

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue  # ignore les lignes vides
                    code = row[0]
                    name = row[1]
                    description = row[2]

                    # Création / Mise à jour
                    Cart.objects.update_or_create(
                        code=code,
                        defaults={
                            "name": name,
                            "description": description
                        }
                    )
            messages.success(request, "Importation réussie.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")

    return redirect("cart-list")

def cart_detail(request, pk):
    cart = Cart.objects.get(pk=pk)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'cart_detail_modal.html', {'cart': cart})

    return redirect('cart-list')

def cart_create(request):
    if request.method == 'POST':
        form = CartForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Panier créé avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = CartForm()
        return render(request, 'cart_form_modal.html', {'form': form, 'type': 'create'})

    return redirect('cart-list')

def cart_update(request, pk):
    cart = Cart.objects.get(pk=pk)
    if request.method == 'POST':
        form = CartForm(request.POST, instance=cart)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Panier mis à jour avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = CartForm(instance=cart)
        return render(request, 'cart_form_modal.html', {'form': form, 'type': 'update', 'cart': cart})

    return redirect('cart-list')

def cart_delete(request, pk):
    cart = Cart.objects.get(pk=pk)
    if request.method == 'POST':
        cart.delete()
        return JsonResponse({'success': True, 'message': 'Panier supprimé avec succès'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'cart_confirm_delete_modal.html', {'cart': cart})

    return redirect('cart-list')

# ====================
# Cart Products Views
# ====================
@login_required
def cartproducts_list_view(request):
    """
    Affiche la liste des produits dans les paniers (CartProducts) avec :
      - Filtrage (ex. search par product name / cart name)
      - Export Excel (si ?export=excel)
      - Pagination
      - Réponse JSON (si AJAX)
    """

    # 1) Paramètres GET
    search = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', 1)
    export_param = request.GET.get('export', '').strip()  # ?export=excel

    # 2) Construction du queryset
    cartprod_qs = CartProducts.objects.select_related('product', 'cart').all()

    # Filtrage / Recherche simple
    if search:
        cartprod_qs = cartprod_qs.filter(
            Q(product__name__icontains=search) |
            Q(cart__name__icontains=search)
        )

    # 3) Export Excel (si ?export=excel)
    if export_param == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "CartProducts"

        # Entêtes de colonnes (selon votre modèle)
        ws.append(["ID", "Produit", "Panier", "Poids", "Date d'ajout", "Date de fin"])

        # Pas de pagination pour l’export
        for cp in cartprod_qs:
            ws.append([
                cp.id,
                cp.product.name if cp.product else "",
                cp.cart.name if cp.cart else "",
                cp.weight if cp.weight else "",
                cp.date_from.strftime("%Y-%m-%d") if cp.date_from else "",
                cp.date_to.strftime("%Y-%m-%d") if cp.date_to else "",
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="cart_products.xlsx"'
        response.write(output.getvalue())
        return response

    # 4) Mode standard (HTML ou AJAX)
    paginator = Paginator(cartprod_qs, 10)  # 10 items par page
    page_obj = paginator.get_page(page_number)

    # Si AJAX => JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            "cart_products": [],
            "pagination": {
                "current_page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        }
        for cp in page_obj:
            data["cart_products"].append({
                "id": cp.id,
                "product": cp.product.name if cp.product else "",
                "cart": cp.cart.name if cp.cart else "",
                "weight": cp.weight,
                "date_from": cp.date_from.isoformat() if cp.date_from else None,
                "date_to": cp.date_to.isoformat() if cp.date_to else None
            })
        return JsonResponse(data, safe=False)

    # Rendu HTML classique
    context = {
        "cart_products": page_obj,
        "search": search,
    }
    # Add empty forms for modals
    context['cartproducts_form'] = CartProductsForm()
    return render(request, "cartproducts_list.html", context)

@login_required
def import_cartproducts_view(request):
    """
    Gère l'import Excel pour CartProducts.
    Suppose qu'on lit un fichier avec colonnes :
      product_code | cart_code | weight | date_from | date_to
    (adapter selon votre vrai fichier)
    """

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect("cartproducts-list")

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active  # ou wb["SheetName"] si nécessaire

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue  # ligne vide

                    product_code = row[0]
                    cart_code = row[1]
                    weight_val = row[2]
                    date_from_val = row[3]
                    date_to_val = row[4]

                    # Récupération du Product
                    product_obj = None
                    if product_code:
                        product_obj = Product.objects.get(code=product_code)

                    # Récupération du Cart
                    cart_obj = None
                    if cart_code:
                        cart_obj = Cart.objects.get(code=cart_code)

                    # Création / mise à jour
                    # Hypothèse : un CartProducts est unique par (product, cart) ?
                    CartProducts.objects.update_or_create(
                        product=product_obj,
                        cart=cart_obj,
                        defaults={
                            "weight": weight_val,
                            "date_from": date_from_val,
                            "date_to": date_to_val
                        }
                    )
            messages.success(request, "Importation réussie.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")

    # Redirection vers la liste
    return redirect("cartproducts-list")

def cartproducts_detail(request, pk):
    cartproducts = CartProducts.objects.get(pk=pk)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'cartproducts_detail_modal.html', {'cartproducts': cartproducts})

    return redirect('cartproducts-list')

def cartproducts_create(request):
    if request.method == 'POST':
        form = CartProductsForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Produit du panier créé avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = CartProductsForm()
        return render(request, 'cartproducts_form_modal.html', {'form': form, 'type': 'create'})

    return redirect('cartproducts-list')

def cartproducts_update(request, pk):
    cartproducts = CartProducts.objects.get(pk=pk)
    if request.method == 'POST':
        form = CartProductsForm(request.POST, instance=cartproducts)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Produit du panier mis à jour avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = CartProductsForm(instance=cartproducts)
        return render(request, 'cartproducts_form_modal.html', {'form': form, 'type': 'update', 'cartproducts': cartproducts})

    return redirect('cartproducts-list')

def cartproducts_delete(request, pk):
    cartproducts = CartProducts.objects.get(pk=pk)
    if request.method == 'POST':
        cartproducts.delete()
        return JsonResponse({'success': True, 'message': 'Produit du panier supprimé avec succès'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'cartproducts_confirm_delete_modal.html', {'cartproducts': cartproducts})

    return redirect('cartproducts-list')

# =========================
# Product Type Views
# =========================
@login_required
def producttype_list_view(request):
    """
    Affiche la liste des types de produits (ProductType) avec :
      - Recherche (code, label, description)
      - Export Excel (via ?export=excel)
      - Pagination
      - Réponse JSON (si AJAX)
    """

    # 1) Paramètres GET
    search = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', 1)
    export_param = request.GET.get('export', '').strip()  # ?export=excel

    # 2) Construction du queryset (filtrage)
    product_types_qs = ProductType.objects.all()

    if search:
        product_types_qs = product_types_qs.filter(
            Q(code__icontains=search) |
            Q(label__icontains=search) |
            Q(description__icontains=search)
        )

    # 3) Export en Excel (si ?export=excel)
    if export_param == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "ProductTypes"

        # En-têtes de colonnes
        ws.append(["ID", "Code", "Label", "Description"])

        # Pas de pagination pour l'export
        for pt in product_types_qs:
            ws.append([
                pt.id,
                pt.code,
                pt.label,
                pt.description
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="product_types.xlsx"'
        response.write(output.getvalue())
        return response

    # 4) Mode standard (HTML ou AJAX)
    paginator = Paginator(product_types_qs, 10)  # 10 types par page
    page_obj = paginator.get_page(page_number)

    # Mode AJAX => JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            "product_types": [],
            "pagination": {
                "current_page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        }
        for pt in page_obj:
            data["product_types"].append({
                "id": pt.id,
                "code": pt.code,
                "label": pt.label,
                "description": pt.description
            })
        return JsonResponse(data, safe=False)

    # Mode HTML
    context = {
        "page_obj": page_obj,
        "search": search
    }
    # Add empty forms for modals
    context['producttype_form'] = ProductTypeForm()
    return render(request, "producttype_list.html", context)
@login_required
def import_producttype_view(request):
    """
    Gère l'import Excel pour ProductType.
    Suppose qu'on lit un fichier avec colonnes :
      code | label | description
      (ligne d'en-tête ignorée, min_row=2)
    """
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect("producttype-list")

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue  # ligne vide
                    code, label, description = row

                    # Mise à jour ou création
                    ProductType.objects.update_or_create(
                        code=code,
                        defaults={
                            "label": label,
                            "description": description,
                        }
                    )
            messages.success(request, "Importation réussie.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")

    return redirect("producttype-list")

def producttype_detail(request, pk):
    producttype = ProductType.objects.get(pk=pk)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'producttype_detail_modal.html', {'producttype': producttype})

    return redirect('producttype-list')

def producttype_create(request):
    if request.method == 'POST':
        form = ProductTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Type de produit créé avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = ProductTypeForm()
        return render(request, 'producttype_form_modal.html', {'form': form, 'type': 'create'})

    return redirect('producttype-list')

def producttype_update(request, pk):
    producttype = ProductType.objects.get(pk=pk)
    if request.method == 'POST':
        form = ProductTypeForm(request.POST, instance=producttype)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Type de produit mis à jour avec succès'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = ProductTypeForm(instance=producttype)
        return render(request, 'producttype_form_modal.html', {'form': form, 'type': 'update', 'producttype': producttype})

    return redirect('producttype-list')

def producttype_delete(request, pk):
    producttype = ProductType.objects.get(pk=pk)
    if request.method == 'POST':
        producttype.delete()
        return JsonResponse({'success': True, 'message': 'Type de produit supprimé avec succès'})

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'producttype_confirm_delete_modal.html', {'producttype': producttype})

    return redirect('producttype-list')

def data_collection_view(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. Create Point of Sale
                commune_code = request.POST.get('commune')
                pos_code = request.POST.get('pointOfSale[code]')
                pos_type = request.POST.get('pointOfSale[type]')
                gps_lat = request.POST.get('pointOfSale[gps_lat]')
                gps_lon = request.POST.get('pointOfSale[gps_lon]')

                # Convert gps_lat and gps_lon to floats, otherwise if they are not a number show an error
                try:
                    gps_lat = float(gps_lat)
                    gps_lon = float(gps_lon)
                except ValueError:
                    messages.error(request, "Invalid Latitude or Longitude.")
                    return redirect("data-collection") # Using redirect for GET

                # Validate that a commune_code was selected
                if not commune_code:
                  messages.error(request, "Please select a commune.")
                  return redirect("data-collection")

                # Check if commune exists
                try:
                  commune = Commune.objects.get(code=commune_code)
                except Commune.DoesNotExist:
                  messages.error(request, f"Commune code {commune_code} is invalid.")
                  return redirect("data-collection")

                # Check if the product of sale code already exists, this should not be the case.
                if PointOfSale.objects.filter(code=pos_code).exists():
                  messages.error(request, f"Point of sale with code {pos_code} already exists.")
                  return redirect("data-collection") # Using redirect for GET

                # Create Point of Sale object
                point_of_sale = PointOfSale.objects.create(
                  code=pos_code,
                  type=pos_type,
                  gps_lat=gps_lat,
                  gps_lon=gps_lon,
                  commune=commune,
                 )
                print("Point of Sale created successfully")  # Debug print

                # 2. Create and Validate Products and Prices
                products_data_json_list = request.POST.getlist('products_json[]')  # Correctly get the list
                print(f"Products JSON list: {products_data_json_list}")  # Debug print

                for product_data_json in products_data_json_list:
                   try:
                      product_data = json.loads(product_data_json)
                      print(f"Product data: {product_data}")  # Debug print

                      product_code = product_data.get('code')
                      product_name = product_data.get('name')
                      product_description = product_data.get('description')
                      product_unit_measure = product_data.get('unit_measure')
                      product_type = product_data.get('type')
                      product_price = product_data.get('price')

                      # Validate Data
                      if not all([product_code, product_name, product_description, product_unit_measure, product_type, product_price]):
                         messages.error(request, "All Product information is required.")
                         return redirect("data-collection")

                      try:
                          product_type_obj = ProductType.objects.get(code=product_type)
                      except ProductType.DoesNotExist:
                         messages.error(request, f"Invalid product type {product_type}.")
                         return redirect("data-collection")

                      try:
                         product_price = Decimal(product_price)
                      except ValueError:
                         messages.error(request, f"Invalid value for price {product_price}.")
                         return redirect("data-collection")

                      # Create product
                      product = Product.objects.create(
                            code=product_code,
                            name=product_name,
                            description=product_description,
                            unit_measure=product_unit_measure,
                            product_type=product_type_obj,
                       )

                      # create product prices
                      ProductPrice.objects.create(
                            product=product,
                            point_of_sale=point_of_sale,
                            value=product_price,
                            date_from=datetime.now().date(),
                            date_to=datetime.now().date() + timedelta(days=365)
                         )
                      print(f"Product {product_name} and price created")  # Debug print
                   except json.JSONDecodeError:
                      messages.error(request, "Invalid JSON data received for a product")
                      return redirect("data-collection")
                   except Exception as e:
                       messages.error(request, f"Error processing product: {e}")
                       return redirect("data-collection")



                messages.success(request, 'Data submitted successfully!')
                return redirect('data-collection')  # Redirect to a success page or back to the form
        except Exception as e:
           messages.error(request, f"An error occurred: {e}")
           return redirect("data-collection")  # Redirect with an error.

    # Render the form on GET request
    communes = Commune.objects.all()
    product_types = ProductType.objects.all()
    context = {
        'communes': communes,
          'product_types': product_types
    }
    return render(request, 'data_collection_form.html', context)


from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import datetime
from .models import ProductPrice, Product, PointOfSale
from django.db.models import Sum, Avg, F, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal

from django.db.models import Sum, Avg, F, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import datetime, date
import calendar



from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from io import BytesIO
from datetime import datetime
from .models import ProductPrice, Product, PointOfSale, CartProducts, ProductType
from django.db.models import Sum, Avg, F, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
import calendar
def generate_report_pdf(request):
    """
    Génère un rapport PDF détaillé sur l'INPC, les produits, les points de vente et les prix.
    """
    buffer = BytesIO()

    # Créer un document PDF avec des marges réduites
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                        leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = [] # List to store the content for the document.

     # Titre principal
    title = Paragraph("Rapport Détaillé INPC", styles['h1'])
    story.append(title)
    story.append(Spacer(1, 12))

      # Data collection for inpc, 6 last months
    today = datetime.now()
    last_6_months = []
    current_year = today.year
    current_month = today.month
    for i in range(6):
        mo = current_month - i
        yr = current_year
        if mo <= 0:
            mo += 12
            yr -= 1
        last_6_months.append((yr, mo))

    inpc_data = []
    for (yr, mo) in reversed(last_6_months):
        inpc_value = calculer_inpc_global_mensuel(annee=yr, mois=mo)
        inpc_data.append({
           'year': yr,
           'month': mo,
           'month_name': calendar.month_name[mo],
           'inpc': round(inpc_value, 2)
       })


     # Create a table for INPC data (last six months)
    subtitle_inpc = Paragraph("INPC des 6 Derniers Mois", styles['h2'])
    story.append(subtitle_inpc)
    story.append(Spacer(1, 6))

    table_data_inpc = [["Mois", "Année", "INPC (%)"]]
    for data in inpc_data:
        table_data_inpc.append([data['month_name'], str(data['year']), str(data['inpc'])])

    table_inpc = Table(table_data_inpc)
    table_inpc.setStyle(TableStyle([
          ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
           ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
          ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
          ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
          ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black) # Add a grid border
    ]))
    story.append(table_inpc)
    story.append(Spacer(1, 12))

    # Table of Product Types
    subtitle_product_types = Paragraph("Types de Produits", styles['h2'])
    story.append(subtitle_product_types)
    story.append(Spacer(1, 6))

    product_type_data = ProductType.objects.all()
    table_data_product_types = [["ID", "Code", "Label", "Description"]]
    for product_type in product_type_data:
        table_data_product_types.append([
            str(product_type.id),
            product_type.code,
            product_type.label,
            product_type.description
        ])

    table_product_types = Table(table_data_product_types)
    table_product_types.setStyle(TableStyle([
         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
          ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black) # Add a grid border
    ]))
    story.append(table_product_types)
    story.append(Spacer(1, 12))

     # Table of Products
    subtitle_products = Paragraph("Liste des Produits", styles['h2'])
    story.append(subtitle_products)
    story.append(Spacer(1, 6))

    product_data = Product.objects.all()
    table_data_products = [["ID", "Code", "Nom", "Description", "Unité de Mesure", "Type de Produit"]]
    for product in product_data:
        table_data_products.append([
            str(product.id),
            product.code,
            product.name,
            product.description,
            product.unit_measure,
            product.product_type.label,
        ])

    table_products = Table(table_data_products)
    table_products.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
         ('GRID', (0, 0), (-1, -1), 1, colors.black)  # Add a grid border
     ]))
    story.append(table_products)
    story.append(Spacer(1, 12))

    # Table of points of sale
    subtitle_pos = Paragraph("Points de Vente", styles['h2'])
    story.append(subtitle_pos)
    story.append(Spacer(1, 6))

    pos_data = PointOfSale.objects.all()
    table_data_pos = [["ID", "Code", "Type", "Latitude GPS", "Longitude GPS", "Commune"]]
    for pos in pos_data:
          table_data_pos.append([
            str(pos.id),
            pos.code,
            pos.type,
            str(pos.gps_lat),
            str(pos.gps_lon),
            pos.commune.name
           ])

    table_pos = Table(table_data_pos)
    table_pos.setStyle(TableStyle([
         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
          ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
           ('GRID', (0, 0), (-1, -1), 1, colors.black) # Add a grid border
     ]))
    story.append(table_pos)
    story.append(Spacer(1, 12))


     # Table of product prices
    subtitle_prices = Paragraph("Prix des Produits", styles['h2'])
    story.append(subtitle_prices)
    story.append(Spacer(1, 6))


    pp_data = ProductPrice.objects.select_related('product','point_of_sale').all()
    table_data_pp = [["ID", "Produit", "Point de Vente", "Prix", "Date Début", "Date Fin"]]

    for price in pp_data:
        date_from = price.date_from.strftime("%Y-%m-%d") if price.date_from else ""
        date_to = price.date_to.strftime("%Y-%m-%d") if price.date_to else ""
        table_data_pp.append([
            str(price.id),
              price.product.name if price.product else "",
              price.point_of_sale.type if price.point_of_sale else "",
               str(price.value),
                date_from,
               date_to,
           ])


    table_pp = Table(table_data_pp)
    table_pp.setStyle(TableStyle([
           ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
           ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
           ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
           ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
           ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
             ('GRID', (0, 0), (-1, -1), 1, colors.black) # Add a grid border
       ]))
    story.append(table_pp)
    story.append(Spacer(1, 12))


    # Build the PDF
    doc.build(story)

    # Get the PDF data
    pdf_value = buffer.getvalue()
    buffer.close()

    # Create the HTTP response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inpc_report.pdf"'
    response.write(pdf_value)

    return response

